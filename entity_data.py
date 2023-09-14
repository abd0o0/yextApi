import json
from openpyxl import load_workbook
import requests


# EntityIds.xlsx should be in the same directory.

def read_excel_rows():
    w = load_workbook('EntityIds.xlsx', data_only=True)
    worksheet = w.active

    rows = worksheet.rows

    headers = [cell.value for cell in next(rows)]

    all_rows = []
    for row in rows:
        data = {}
        for title, cell in zip(headers, row):
            data[title] = cell.value

        all_rows.append(data)
    w.close()

    return all_rows


def fetch_data(entity_id, sheet, start_date="2019-10-01"):
    url = "https://api.yext.com/v2/accounts/me/analytics/reports?api_key=8591b0ac53f3cf563cc88b7f674bb856&v=20201010"

    payload = {
        "metrics": ['GOOGLE_SEARCH_VIEWS', 'GOOGLE_MAP_VIEWS'],
        "dimensions": ["DAYS"],
        "filters": {
            "startDate": start_date,
            "entityIds": [entity_id],
        }
    }
    js = json.dumps(payload)
    headers = {
        'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=js).json()
    data = response['response']['data']
    col_index = 3
    for _ in range(len(data)):

        day_cell = sheet.cell(row=1, column=col_index)
        day_cell.value = 'day'
        col_index += 1

        gmv_cell = sheet.cell(row=1, column=col_index)
        gmv_cell.value = 'Google Maps Views'
        col_index += 1

        gsv_cell = sheet.cell(row=1, column=col_index)
        gsv_cell.value = 'Google Search Views'
        col_index += 1
    return data


def main():
    wb = load_workbook('EntityIds.xlsx')
    sheet = wb.active

    rows = read_excel_rows()
    starting_date = input('starting date (YYYY-MM-DD): ')
    row_index = 2
    print(f"{len(rows)} Entities found")
    for row in rows:
        col_index = 3

        ids = row['Entity ID']
        name = row['name']
        print(f"{row_index - 1} - Entity-ID {ids} - name {name}.")
        # name = row['name']
        data = fetch_data(ids, sheet, starting_date)
        print(data)
        break
        #for day_data in data:

            #day_cell = sheet.cell(row=row_index, column=col_index)
            #day_cell.value = day_data['day']
            #col_index += 1
#
            #gmv_cell = sheet.cell(row=row_index, column=col_index)
            #gmv_cell.value = day_data['Google Maps Views']
            #col_index += 1
#
            #gsv_cell = sheet.cell(row=row_index, column=col_index)
            #gsv_cell.value = day_data['Google Search Views']
            #col_index += 1
#
            #wb.save('Entities-data.xlsx')

        #row_index += 1
    #wb.close()
    print("done!")
    return


if __name__ == '__main__':
    main()
