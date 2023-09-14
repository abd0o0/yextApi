from openpyxl import Workbook
import requests

wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Entity ID'
sheet.cell(row=1, column=2).value = 'name'
#sheet.cell(row=1, column=3 ).value= 'googlePlaceID'
index = 2
pageToken = ''
temp = ''

cond = True
while cond:
    url = f"https://api.yext.com/v2/accounts/me/entities?api_key=8591b0ac53f3cf563cc88b7f674bb856&v=20201010&limit=50{temp}{pageToken}"
    response = requests.request("GET", url)
    js = response.json()
    resp = js['response']
    keys = resp.keys()
    entities = resp['entities']
    for i in entities:
        ids = i['meta']['id']
        name = i['name']

      # try:
      #     google = i['googlePlaceId']
      #     sheet.cell(row=index, column=3).value = google
      # except:
      #     print('notfound')

        sheet.cell(row=index, column=1).value = ids
        sheet.cell(row=index, column=2).value = name
        wb.save('EntityIds.xlsx')
        index += 1

    if 'pageToken' in keys:
        pageToken = resp['pageToken']
        print(pageToken)
        temp = '&pageToken='
    else:
        cond = False
wb.close()
